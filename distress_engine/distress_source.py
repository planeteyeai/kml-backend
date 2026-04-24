import cv2
import io
import math
import numpy as np
import os
import pandas as pd
import re
import tempfile
import uuid
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


HEADER_PATTERN = re.compile(r"H_?(\d+)")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
KML_BACKEND_DIR = os.path.normpath(os.path.join(BASE_DIR, ".."))
DEFAULT_GEO_EXCEL_DIR = os.getenv(
    "CHAINAGE_GEO_EXCEL_DIR",
    os.path.join(KML_BACKEND_DIR, "pipeline", "Excels"),
)
SIDE_GEO_FILES = {
    "LHS": "LHS_L1.xlsx",
    "RHS": "RHS_L1.xlsx",
}
CHAINAGE_GEO_CACHE = {"LHS": None, "RHS": None}


def rotate_rectangle(img):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
    mask = cv2.inRange(hsv, np.array([40, 40, 40]), np.array([80, 255, 255]))
    if np.sum(mask) == 0:
        _, mask = cv2.threshold(gray, 240, 255, cv2.THRESH_BINARY_INV)
    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    if not contours:
        return img
    cnt = max(contours, key=cv2.contourArea)
    rect = cv2.minAreaRect(cnt)
    (w, h) = rect[1]
    angle = rect[2]
    angle = angle + 90 if w < h else angle
    M = cv2.getRotationMatrix2D((img.shape[1] / 2, img.shape[0] / 2), angle, 1)
    return cv2.warpAffine(img, M, (img.shape[1], img.shape[0]), borderValue=(255, 255, 255))


def force_horizontal(img):
    if img.shape[0] > img.shape[1]:
        img = cv2.rotate(img, cv2.ROTATE_90_CLOCKWISE)
    return img


def pca_correct(img):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    _, thresh = cv2.threshold(gray, 10, 255, cv2.THRESH_BINARY)
    contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    if not contours:
        return img
    cnt = max(contours, key=cv2.contourArea)
    data_pts = np.squeeze(cnt).astype(np.float32)
    _, eigenvectors = cv2.PCACompute(data_pts, mean=None)
    angle = np.degrees(np.arctan2(eigenvectors[0, 1], eigenvectors[0, 0]))
    center = (img.shape[1] // 2, img.shape[0] // 2)
    M = cv2.getRotationMatrix2D(center, angle, 1.0)
    return cv2.warpAffine(img, M, (img.shape[1], img.shape[0]))


def fix_direction(img):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    if np.mean(gray[:, :10]) > np.mean(gray[:, -10:]):
        img = cv2.flip(img, 1)
    if np.mean(gray[:10, :]) > np.mean(gray[-10:, :]):
        img = cv2.flip(img, 0)
    return img


def create_temp_dataframe(img, filename):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    h, w = gray.shape
    total_length_m = 100
    target_dist = np.arange(0, total_length_m + 0.05, 0.05)
    pixel_to_meter_w = total_length_m / w
    target_pix_w = (target_dist / pixel_to_meter_w).astype(int)
    target_pix_w[target_pix_w >= w] = w - 1

    spacing_m = 0.05
    pixel_to_meter_h = total_length_m / h
    step_pixels_h = spacing_m / pixel_to_meter_h
    h_pos = np.arange(0, h, step_pixels_h)
    h_pos = np.unique(np.round(h_pos).astype(int))
    h_pos = h_pos[h_pos < h]

    data = {}
    for i, y in enumerate(h_pos):
        data[f"H_{i + 1}"] = gray[y, target_pix_w].astype(float)

    df = pd.DataFrame(data)
    df.insert(0, "Distance_meter", target_dist)

    mask = df.filter(like="H_").notna().any(axis=1)
    df = df[mask].copy()
    df["Distance_meter"] -= df["Distance_meter"].min()

    cols_to_keep = ["Distance_meter"]
    for col in df.columns:
        if col != "Distance_meter" and (df[col] > 0).any():
            cols_to_keep.append(col)
    df = df[cols_to_keep]

    h_cols = [c for c in df.columns if c.startswith("H_")]
    df.columns = ["Distance_meter"] + [f"H_{i + 1}" for i in range(len(h_cols))]
    h_cols = [c for c in df.columns if c.startswith("H_")]
    total_cols = len(h_cols)

    start_idx, end_idx = 0, total_cols
    if "LHS" in filename:
        start_idx = int(0.15 * total_cols)
        end_idx = total_cols - int(0.20 * total_cols)
    elif "RHS" in filename:
        start_idx = int(0.20 * total_cols)
        end_idx = total_cols - int(0.15 * total_cols)

    df = df[["Distance_meter"] + h_cols[start_idx:end_idx]]
    h_cols = [c for c in df.columns if c.startswith("H_")]
    df.columns = ["Distance_meter"] + [f"H_{i + 1}" for i in range(len(h_cols))]
    return df


def generate_expanded_excel(temp_path, output_path):
    wb_in = load_workbook(temp_path)
    ws_in = wb_in.active
    source_headers = [ws_in.cell(1, c).value for c in range(1, ws_in.max_column + 1)]

    source_h_columns = []
    output_headers = ["Distance_meter"]
    for i, h in enumerate(source_headers, start=1):
        if isinstance(h, str) and HEADER_PATTERN.fullmatch(h):
            suffix = HEADER_PATTERN.fullmatch(h).group(1)
            source_h_columns.append((i, h, suffix))
            output_headers.extend([h, f"max_H{suffix}", f"min_H{suffix}", f"AvgH{suffix}", f"diffH{suffix}", f"newH{suffix}"])

    wb_out = Workbook()
    ws_out = wb_out.active
    for col_idx, header in enumerate(output_headers, start=1):
        ws_out.cell(1, col_idx, header)

    max_row = ws_in.max_row
    for row in range(2, max_row + 1):
        ws_out.cell(row, 1, ws_in.cell(row, 1).value)
        out_col = 2
        for source_col, _, _ in source_h_columns:
            h_letter = get_column_letter(out_col)
            max_letter = get_column_letter(out_col + 1)
            min_letter = get_column_letter(out_col + 2)
            avg_letter = get_column_letter(out_col + 3)
            diff_letter = get_column_letter(out_col + 4)

            val = ws_in.cell(row, source_col).value
            ws_out.cell(row, out_col, val)
            ws_out.cell(row, out_col + 1, f"=MAX(${h_letter}$2:${h_letter}${max_row})-{h_letter}{row}")
            ws_out.cell(row, out_col + 2, f"=MIN(${h_letter}$2:${h_letter}${max_row})-{h_letter}{row}")
            ws_out.cell(row, out_col + 3, f"=AVERAGE({max_letter}{row},{min_letter}{row})")
            ws_out.cell(row, out_col + 4, f"={h_letter}{row}-{avg_letter}{row}")
            if row < max_row:
                ws_out.cell(row, out_col + 5, f"={diff_letter}{row + 1}-{diff_letter}{row}")
            out_col += 6

    for col in range(2, ws_out.max_column + 1):
        header = ws_out.cell(1, col).value
        if not header.startswith("newH"):
            ws_out.column_dimensions[get_column_letter(col)].hidden = True
    wb_out.save(output_path)


def normalize_grayscale(matrix):
    min_val = 0.0
    for row in matrix:
        for v in row:
            if v < min_val:
                min_val = v
    range_val = abs(min_val) if min_val != 0 else 1.0
    out = []
    for row in matrix:
        out.append([max(0, min(255, int(round(255 * (1 + (v / range_val)))))) for v in row])
    return out


def connected_components(mask):
    rows = len(mask)
    cols = len(mask[0]) if rows else 0
    labels = [[0] * cols for _ in range(rows)]
    label = 1
    components = {}
    for i in range(rows):
        for j in range(cols):
            if mask[i][j] == 1 and labels[i][j] == 0:
                queue = [(i, j)]
                labels[i][j] = label
                pixels = []
                while queue:
                    r, c = queue.pop(0)
                    pixels.append((r, c))
                    for dr, dc in [(-1, 0), (1, 0), (0, -1), (0, 1)]:
                        nr, nc = r + dr, c + dc
                        if 0 <= nr < rows and 0 <= nc < cols and mask[nr][nc] == 1 and labels[nr][nc] == 0:
                            labels[nr][nc] = label
                            queue.append((nr, nc))
                components[label] = pixels
                label += 1
    return components


def compute_gradient_edges(matrix, threshold=20):
    rows = len(matrix)
    cols = len(matrix[0]) if rows else 0
    edge_count = 0
    for i in range(rows):
        for j in range(cols):
            dx = abs(matrix[i][j + 1] - matrix[i][j]) if j < cols - 1 else 0
            dy = abs(matrix[i + 1][j] - matrix[i][j]) if i < rows - 1 else 0
            mag = math.sqrt(dx * dx + dy * dy)
            if mag >= threshold:
                edge_count += 1
    return edge_count


def calc_confidence(depth, length, sensors):
    d_conf = min(1.0, abs(depth) / 80.0)
    l_conf = min(1.0, length / 1.0)
    s_conf = min(1.0, sensors / 6.0)
    return round(min(1.0, 0.45 * d_conf + 0.3 * l_conf + 0.25 * s_conf), 2)


def compute_newh_matrix(df):
    h_cols = [c for c in df.columns if isinstance(c, str) and c.startswith("H_")]
    if not h_cols:
        raise ValueError("No H_ columns found in generated Excel")

    distances = pd.to_numeric(df["Distance_meter"], errors="coerce").fillna(0.0).astype(float).tolist()
    matrix = [[0.0] * len(h_cols) for _ in range(len(df))]

    for j, h_col in enumerate(h_cols):
        series = pd.to_numeric(df[h_col], errors="coerce").fillna(0.0).astype(float).tolist()
        col_max = max(series) if series else 0.0
        col_min = min(series) if series else 0.0
        diff_vals = []
        for val in series:
            max_h = col_max - val
            min_h = col_min - val
            avg_h = (max_h + min_h) / 2.0
            diff_h = val - avg_h
            diff_vals.append(diff_h)
        for i in range(len(series)):
            matrix[i][j] = float(diff_vals[i + 1] - diff_vals[i] if i < len(series) - 1 else 0.0)
    return distances, matrix


def compute_newh_matrix_from_df(df):
    return compute_newh_matrix(df)


def cluster_rows(row_arr, gap, defect_type):
    if not row_arr:
        return []
    sorted_rows = sorted(row_arr, key=lambda x: x["dist"])
    clusters = []
    s = sorted_rows[0]
    e = sorted_rows[0]
    max_d = sorted_rows[0]["minH"]
    max_s = sorted_rows[0]["numNeg"]
    max_lat = sorted_rows[0]["latWidth"]
    acc_rows = [sorted_rows[0]]
    step = sorted_rows[1]["dist"] - sorted_rows[0]["dist"] if len(sorted_rows) > 1 else 0.05

    def flush():
        length = round((e["dist"] - s["dist"] + step), 3)
        if max_d <= -60:
            sev = "SEVERE"
        elif max_d <= -30:
            sev = "MODERATE"
        else:
            sev = "MILD"
        clusters.append(
            {
                "type": defect_type,
                "start": round(s["dist"], 3),
                "end": round(e["dist"], 3),
                "length": length,
                "maxDepth": max_d,
                "maxSensors": max_s,
                "maxLatWidth": round(max_lat, 2),
                "severity": sev,
                "confidence": calc_confidence(max_d, length, max_s),
                "depths": [v for r in acc_rows for v in r["vals"]],
            }
        )

    for r in sorted_rows[1:]:
        if r["dist"] - e["dist"] <= gap:
            e = r
            max_d = min(max_d, r["minH"])
            max_s = max(max_s, r["numNeg"])
            max_lat = max(max_lat, r["latWidth"])
            acc_rows.append(r)
        else:
            flush()
            s = r
            e = r
            max_d = r["minH"]
            max_s = r["numNeg"]
            max_lat = r["latWidth"]
            acc_rows = [r]
    flush()
    return clusters


def detect_distresses_from_generated_excel(temp_df):
    distances, matrix = compute_newh_matrix_from_df(temp_df)
    if not matrix or not matrix[0]:
        return {"defects": [], "meta": {"components": 0, "edges": 0}}

    gs_thresh = 180
    min_blob = 3
    p_depth, p_sensors, p_gap = -130, 4, 0.5
    c_depth, c_sensors, c_gap = -15, 2, 0.3
    a_sensors, a_max_depth, a_gap = 5, -50, 0.5
    lateral_step = 0.1

    gray_matrix = normalize_grayscale(matrix)
    mask = [[1 if v < gs_thresh else 0 for v in row] for row in gray_matrix]
    components = connected_components(mask)
    edge_count = compute_gradient_edges(matrix, threshold=20)
    component_count = sum(1 for px in components.values() if len(px) >= min_blob)

    rows = []
    for i, row in enumerate(matrix):
        neg_vals = [v for v in row if v < 0]
        min_h = min(neg_vals) if neg_vals else 0.0
        num_below_10 = sum(1 for v in row if v <= -10)
        num_below_5 = sum(1 for v in row if v <= -5)
        num_neg = len(neg_vals)
        idxs = [j for j, v in enumerate(row) if v <= -10]
        lat_width = ((max(idxs) - min(idxs)) * lateral_step) if idxs else 0.0
        rows.append(
            {
                "dist": float(distances[i]) if i < len(distances) else round(i * 0.05, 3),
                "minH": float(min_h),
                "numBelow10": int(num_below_10),
                "numBelow5": int(num_below_5),
                "numNeg": int(num_neg),
                "latWidth": float(lat_width),
                "vals": [float(v) for v in row],
            }
        )

    pothole_rows = [r for r in rows if r["minH"] <= p_depth and r["numBelow10"] >= p_sensors]
    pothole_set = {r["dist"] for r in pothole_rows}
    crack_rows = [
        r
        for r in rows
        if r["minH"] <= c_depth and 1 <= r["numBelow10"] <= c_sensors and r["dist"] not in pothole_set
    ]
    crack_set = {r["dist"] for r in crack_rows}
    alligator_rows = [
        r
        for r in rows
        if r["numBelow5"] >= a_sensors
        and r["minH"] >= a_max_depth
        and r["dist"] not in pothole_set
        and r["dist"] not in crack_set
    ]

    potholes = cluster_rows(pothole_rows, p_gap, "pothole")
    cracks = cluster_rows(crack_rows, c_gap, "crack")
    alligator = cluster_rows(alligator_rows, a_gap, "alligator")

    all_defects = []
    for defect in potholes + cracks + alligator:
        if defect["type"] != "pothole":
            all_defects.append(defect)
            continue

        depths = [v for v in defect.get("depths", []) if v < 0]
        if len(depths) <= 5:
            all_defects.append(defect)
            continue

        deepest = min(depths)
        cutoff = deepest + 10
        pothole_depths = [v for v in depths if v <= cutoff]
        crack_depths = [v for v in depths if v > cutoff]
        if not crack_depths:
            all_defects.append(defect)
            continue

        pothole_def = dict(defect)
        pothole_def["type"] = "pothole"
        pothole_def["maxDepth"] = min(pothole_depths)
        crack_def = dict(defect)
        crack_def["type"] = "crack"
        crack_def["maxDepth"] = min(crack_depths)
        all_defects.extend([pothole_def, crack_def])

    all_defects = sorted(all_defects, key=lambda d: d["start"])
    pothole_defs = [d for d in all_defects if d["type"] == "pothole"]
    if len(pothole_defs) >= 5:
        sorted_potholes = sorted(pothole_defs, key=lambda d: d["maxDepth"])
        keep_potholes = sorted_potholes[:2]
        for d in all_defects:
            if d["type"] == "pothole" and d not in keep_potholes:
                d["type"] = "crack"

    pothole_bucket = [d for d in all_defects if d["type"] == "pothole"]
    crack_bucket = [d for d in all_defects if d["type"] == "crack"]
    alligator_bucket = [d for d in all_defects if d["type"] == "alligator"]
    ordered_defects = pothole_bucket + crack_bucket + alligator_bucket

    output = []
    for idx, d in enumerate(ordered_defects, start=1):
        dtype = d["type"]
        if dtype == "pothole":
            final_type = "reported_pothole" if d["severity"] == "SEVERE" else "predicted_pothole"
        elif dtype == "crack":
            final_type = "reported_crack" if d["severity"] == "SEVERE" else "predicted_crack"
        elif dtype == "alligator":
            if d["severity"] == "SEVERE":
                final_type = "reported_alligator_crack"
            elif d["severity"] == "MODERATE":
                final_type = "predicted_alligator_crack"
            else:
                continue
        else:
            continue
        output.append(
            {
                "id": idx,
                "type": final_type,
                "severity": d["severity"],
                "start": round(float(d["start"]), 3),
                "end": round(float(d["end"]), 3),
                "length": round(float(d["length"]), 3),
                "max_depth": int(round(float(d["maxDepth"]))),
                "width": round(float(d["maxLatWidth"]), 2),
                "sensors": int(d["maxSensors"]),
                "confidence": round(float(d["confidence"]), 2),
            }
        )

    counts = {
        "reported_crack": 0,
        "predicted_crack": 0,
        "reported_pothole": 0,
        "predicted_pothole": 0,
        "reported_alligator_crack": 0,
        "predicted_alligator_crack": 0,
    }
    for defect in output:
        d_type = defect.get("type")
        if d_type in counts:
            counts[d_type] += 1
    return {"defects": output, "counts": counts, "meta": {"components": int(component_count), "edges": int(edge_count)}}


def trim_temp_dataframe_rows(df, top_rows=3, bottom_rows=3):
    total_trim = top_rows + bottom_rows
    if len(df) > total_trim:
        return df.iloc[top_rows : len(df) - bottom_rows].reset_index(drop=True)
    return df.iloc[0:0].copy()


def extract_chainage_start_meters(filename):
    match = re.search(r"Chainage_(\d+(?:\.\d+)?)_to_(\d+(?:\.\d+)?)", filename or "", flags=re.IGNORECASE)
    if not match:
        raise ValueError("Invalid filename. Expected pattern like: Chainage_0.200_to_0.300_LHS_L1.png")
    start_km = float(match.group(1))
    end_km = float(match.group(2))
    if end_km <= start_km:
        raise ValueError("Invalid chainage range in filename: end must be greater than start.")
    return start_km * 1000.0


def infer_side_from_filename(filename):
    name = (filename or "").upper()
    if "LHS" in name:
        return "LHS"
    if "RHS" in name:
        return "RHS"
    return None


def load_chainage_geo_df(side):
    side = (side or "").upper()
    if side not in SIDE_GEO_FILES:
        return None
    cached = CHAINAGE_GEO_CACHE.get(side)
    if cached is not None:
        return cached

    excel_path = os.path.join(DEFAULT_GEO_EXCEL_DIR, SIDE_GEO_FILES[side])
    if not os.path.exists(excel_path):
        CHAINAGE_GEO_CACHE[side] = pd.DataFrame()
        return CHAINAGE_GEO_CACHE[side]

    df = pd.read_excel(excel_path)
    needed = ["Chainage Start", "Chainage End", "Latitude", "Longitude"]
    if not set(needed).issubset(df.columns):
        CHAINAGE_GEO_CACHE[side] = pd.DataFrame()
        return CHAINAGE_GEO_CACHE[side]

    df = df[needed].copy()
    for col in needed:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df = df.dropna(subset=needed).sort_values("Chainage Start").reset_index(drop=True)
    CHAINAGE_GEO_CACHE[side] = df
    return df


def get_latlon_for_chainage(df_geo, chainage_m):
    if df_geo is None or df_geo.empty:
        return None, None, None
    chainage_km = float(chainage_m) / 1000.0
    exact = df_geo[(df_geo["Chainage Start"] <= chainage_km) & (chainage_km < df_geo["Chainage End"])]
    if not exact.empty:
        row = exact.iloc[0]
        return float(row["Latitude"]), float(row["Longitude"]), float(row["Chainage Start"])
    nearest_idx = (df_geo["Chainage Start"] - chainage_km).abs().idxmin()
    row = df_geo.loc[nearest_idx]
    return float(row["Latitude"]), float(row["Longitude"]), float(row["Chainage Start"])


def attach_chainage_geo_to_defects(result, filename):
    side = infer_side_from_filename(filename)
    df_geo = load_chainage_geo_df(side) if side else pd.DataFrame()
    for defect in result.get("defects", []):
        start_m = float(defect.get("start", 0.0) or 0.0)
        end_m = float(defect.get("end", 0.0) or 0.0)
        middle_m = (start_m + end_m) / 2.0
        lat, lon, chainage_start_km = get_latlon_for_chainage(df_geo, middle_m)
        defect["side"] = side
        defect["latitude"] = round(lat, 8) if lat is not None else None
        defect["longitude"] = round(lon, 8) if lon is not None else None
        defect["matched_chainage_start_km"] = round(chainage_start_km, 3) if chainage_start_km is not None else None


def is_99x_mod_hundred(km_value):
    meters = float(km_value) * 1000.0
    rem = meters % 100.0
    return 99.0 < rem < 100.0


def process_image_bytes(image_bytes, filename, process_type, persist_excels=True):
    file_id = str(uuid.uuid4())
    nparr = np.frombuffer(image_bytes, np.uint8)
    img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    if img is None:
        raise ValueError("Invalid image")

    if process_type == "already_rotated":
        processed = img
    elif process_type == "down_to_up":
        processed = fix_direction(pca_correct(force_horizontal(rotate_rectangle(img))))
    elif process_type == "up_to_down":
        processed = fix_direction(pca_correct(force_horizontal(rotate_rectangle(cv2.flip(img, 0)))))
    else:
        raise ValueError("Invalid process_type")

    temp_df = create_temp_dataframe(processed, filename)
    temp_df = trim_temp_dataframe_rows(temp_df, top_rows=3, bottom_rows=3)
    result = detect_distresses_from_generated_excel(temp_df=temp_df)
    base_chainage_m = extract_chainage_start_meters(filename)

    for defect in result.get("defects", []):
        defect["start"] = round(float(defect.get("start", 0.0) or 0.0) + base_chainage_m, 3)
        defect["end"] = round(float(defect.get("end", 0.0) or 0.0) + base_chainage_m, 3)
        length = float(defect.get("length", 0.0) or 0.0)
        depth = float(defect.get("max_depth", 0.0) or 0.0)

        defect["reported_depth"] = None
        defect["reported_width"] = None
        defect["total_width"] = None
        defect["pothole_area"] = None
        if defect.get("type") == "reported_crack":
            reported_depth = (depth / 10.0) * (10 ** 3)
            reported_width = reported_depth * 0.77
            defect["reported_depth"] = round(reported_depth, 3)
            defect["reported_width"] = round(reported_width, 3)
            defect["total_width"] = round(length * reported_width, 3)
        elif defect.get("type") == "reported_pothole":
            defect["pothole_area"] = round(3.14 * ((length / 2.0) ** 2), 3)

    attach_chainage_geo_to_defects(result, filename)
    for defect in result.get("defects", []):
        defect["start"] = round(float(defect.get("start", 0.0) or 0.0) / 1000.0, 4)
        defect["end"] = round(float(defect.get("end", 0.0) or 0.0) / 1000.0, 4)

    result["defects"] = [
        defect
        for defect in result.get("defects", [])
        if not (
            is_99x_mod_hundred(defect.get("start", 0.0) or 0.0)
            or is_99x_mod_hundred(defect.get("end", 0.0) or 0.0)
        )
    ]

    if persist_excels:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = os.path.join(temp_dir, file_id + "_raw.xlsx")
            expanded_path = os.path.join(temp_dir, file_id + "_expanded.xlsx")
            temp_df.to_excel(temp_path, index=False)
            generate_expanded_excel(temp_path, expanded_path)

    result["run_id"] = file_id
    result["download_endpoints"] = {
        "raw_excel": "/download-excel/raw",
        "expanded_excel": "/download-excel/expanded",
        "expanded_excel_data": "/expanded-excel-data",
    }
    return file_id, result, None, None


def process_rotated_image_job(image_name, image_bytes):
    _, result, _, _ = process_image_bytes(image_bytes, image_name, "already_rotated", persist_excels=False)
    return image_name, result
